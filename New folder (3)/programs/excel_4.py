from openpyxl import load_workbook

wb=load_workbook("second_file.xlsx")

sheet=wb.active

data=((1,2,3),(4,5,6),(7,8,9),(10,11,12))

for i in data:
    sheet.append(i)

for i in sheet.iter_rows(min_row=1,min_col=1,max_row=4,max_col=3):
    for j in i:
        print(j.value,end=" ")
    print()

wb.save("third_file.xlsx")