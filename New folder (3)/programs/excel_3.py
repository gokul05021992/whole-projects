from openpyxl import load_workbook

wb=load_workbook("second_file.xlsx")

sheet=wb.active

a=sheet['A1']

b=sheet['A2']

c=sheet.cell(row=3,column=3)

print("A1=",a.value)
print("A2=",b.value)
print("r3-c3=",c.value)